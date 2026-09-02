from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, Body, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl.styles import PatternFill

APP_ROOT = Path(__file__).resolve().parents[2]
EXPORT_DIR = APP_ROOT / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(APP_ROOT / "db"))
sys.path.insert(0, str(APP_ROOT / "backend"))
import db_manager as dbm  # noqa: E402
import auth  # noqa: E402
import matched_review_adapter as review  # noqa: E402

router = APIRouter(dependencies=[Depends(auth.require_user)])

# Soft red fill for rows the user has flagged as mismatches on matched_records.
_FLAGGED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")


def _check_table_access(table: str, user: dict) -> None:
    allowed = dbm.get_user_allowed_tables(user["user_id"], user["role"])
    if table not in allowed:
        raise HTTPException(status_code=403, detail=f"You don't have access to '{table}'")


def _filtered_frame(table: str, filters: dict | None, start_date: str | None, end_date: str | None,
                    user: dict | None = None):
    # Date filtering is delegated to dbm.fetch_table's norm_date()-based
    # SQL comparison (same logic the Evaluation page uses) rather than
    # re-parsing dates here with pandas: pd.to_datetime() defaults to
    # month-first parsing, which silently misreads a day-first
    # dd/mm/yyyy date whenever the day is <= 12 — exporting used to
    # disagree with what the Evaluation page filtered to for exactly
    # that reason.
    if user and user.get("role") == "user" and table == "matched_records":
        rows = review.fetch_review(
            filters=filters, limit=200000, offset=0,
            start_date=start_date, end_date=end_date,
        )
        return pd.DataFrame(rows, columns=review.REVIEW_COLUMNS)
    rows = dbm.fetch_table(
        table, filters=filters, limit=200000,
        start_date=start_date, end_date=end_date,
    )
    return pd.DataFrame(rows)


def _highlight_flagged_mismatches(ws, df: pd.DataFrame) -> None:
    """Paint entire Excel rows red when user_flagged_mismatch == 1."""
    if "user_flagged_mismatch" not in df.columns or df.empty:
        return
    for i, flagged in enumerate(df["user_flagged_mismatch"].tolist()):
        if flagged in (1, True, "1"):
            # openpyxl rows are 1-indexed; row 1 is the header, so data starts at 2
            excel_row = i + 2
            for cell in ws[excel_row]:
                cell.fill = _FLAGGED_FILL


@router.post("/table/{table}")
def export_table(table: str, payload: dict = Body(default={}), user=Depends(auth.require_user)):
    _check_table_access(table, user)
    filters = payload.get("filters")
    start_date = payload.get("start_date")
    end_date = payload.get("end_date")

    df = _filtered_frame(table, filters, start_date, end_date, user=user)
    fname = f"{table}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = EXPORT_DIR / fname
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        sheet = table[:31]
        df.to_excel(writer, index=False, sheet_name=sheet)
        if table == "matched_records":
            _highlight_flagged_mismatches(writer.sheets[sheet], df)
    return FileResponse(path, filename=fname,
                         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/all")
def export_all(payload: dict = Body(default={}), user=Depends(auth.require_user)):
    start_date = payload.get("start_date")
    end_date = payload.get("end_date")
    allowed_tables = dbm.get_user_allowed_tables(user["user_id"], user["role"])

    fname = f"database_history_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = EXPORT_DIR / fname
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for table in allowed_tables:
            df = _filtered_frame(table, None, start_date, end_date, user=user)
            if df.empty:
                cols = (review.REVIEW_COLUMNS if user.get("role") == "user" and table == "matched_records"
                        else dbm.table_columns(table))
                df = pd.DataFrame(columns=cols)
            sheet = table[:31]
            df.to_excel(writer, index=False, sheet_name=sheet)
            if table == "matched_records":
                _highlight_flagged_mismatches(writer.sheets[sheet], df)
    return FileResponse(path, filename=fname,
                         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")